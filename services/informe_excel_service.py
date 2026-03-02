import logging
from io import BytesIO
from datetime import datetime

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

COLOR_CABECERA = "1A3D6E"
COLOR_FILA_IMPAR = "E8F0FB"
COLOR_FILA_PAR = "FFFFFF"
COLOR_EXITO = "D4EDDA"
COLOR_ALERTA = "FFE5B4"
COLOR_ERROR = "F8D7DA"

MESES_ES = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
    5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
    9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
}

COLUMNAS = [
    ("Empleado", 30),
    ("Jornada (h/sem)", 14),
    ("Laborables", 12),
    ("Fichados", 11),
    ("Errores", 10),
    ("Sin fichar", 12),
    ("Horas reales", 13),
    ("Objetivo (h)", 13),
    ("Diferencia (h)", 14),
    ("Horas extra (h)", 14),
]

CAMPOS = [
    "nombre", "jornada", "laborables", "fichados",
    "errores", "sin_fichar", "horas_reales", "objetivo",
    "diferencia", "horas_extra",
]

THIN = Side(style="thin", color="CCCCCC")
BORDE = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _estilo_cabecera():
    return Font(name="Calibri", bold=True, color="FFFFFF", size=10)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _escribir_cabecera(ws, fila: int = 1):
    for col_idx, (nombre, ancho) in enumerate(COLUMNAS, start=1):
        cell = ws.cell(row=fila, column=col_idx, value=nombre)
        cell.font = _estilo_cabecera()
        cell.fill = _fill(COLOR_CABECERA)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDE
        ws.column_dimensions[get_column_letter(col_idx)].width = ancho
    ws.row_dimensions[fila].height = 30


def _escribir_fila_datos(ws, fila: int, d: dict, par: bool):
    color_base = COLOR_FILA_PAR if par else COLOR_FILA_IMPAR
    for col_idx, campo in enumerate(CAMPOS, start=1):
        valor = d[campo]
        cell = ws.cell(row=fila, column=col_idx, value=valor)
        fill_color = color_base
        if campo == "sin_fichar":
            if d["sin_fichar"] == 0:
                fill_color = COLOR_EXITO
            elif d["sin_fichar"] < 3:
                fill_color = COLOR_ALERTA
            else:
                fill_color = COLOR_ERROR
        cell.fill = _fill(fill_color)
        cell.border = BORDE
        cell.alignment = Alignment(
            horizontal="center" if col_idx > 1 else "left",
            vertical="center",
        )
        cell.font = Font(name="Calibri", size=9)
    ws.row_dimensions[fila].height = 18


def _agregar_metadata(ws, mes: int, anno: int, total: int):
    ws["A1"] = "WorkTimeAsistem PRODE — Fundación PRODE"
    ws["A1"].font = Font(name="Calibri", bold=True, size=12, color=COLOR_CABECERA)
    nombre_mes = MESES_ES.get(mes, str(mes))
    ws["A2"] = f"Informe de Control Horario — {nombre_mes} {anno}"
    ws["A2"].font = Font(name="Calibri", size=10, color="2E6DA4")
    ws["A3"] = f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')} · {total} empleados"
    ws["A3"].font = Font(name="Calibri", size=8, color="888888", italic=True)


class InformeExcelService:
    def generar_excel_global(self, lista_data: list[dict], mes: int, anno: int) -> bytes:
        wb = openpyxl.Workbook()
        ws = wb.active
        nombre_mes = MESES_ES.get(mes, str(mes))
        ws.title = f"{nombre_mes} {anno}"

        _agregar_metadata(ws, mes, anno, len(lista_data))
        ws.append([])

        FILA_INICIO = 5
        _escribir_cabecera(ws, fila=FILA_INICIO)

        for i, d in enumerate(lista_data):
            _escribir_fila_datos(ws, FILA_INICIO + 1 + i, d, par=(i % 2 == 0))

        ws.freeze_panes = f"A{FILA_INICIO + 1}"
        ws.auto_filter.ref = (
            f"A{FILA_INICIO}:{get_column_letter(len(COLUMNAS))}{FILA_INICIO + len(lista_data)}"
        )

        buffer = BytesIO()
        wb.save(buffer)
        logger.info("Excel global generado: %d empleados", len(lista_data))
        return buffer.getvalue()

    def generar_excel_individual(self, emp_data: dict, mes: int, anno: int) -> bytes:
        wb = openpyxl.Workbook()
        ws = wb.active
        nombre_mes = MESES_ES.get(mes, str(mes))
        ws.title = f"{emp_data['nombre'][:20]}"

        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 20

        titulo_font = Font(name="Calibri", bold=True, size=12, color=COLOR_CABECERA)
        ws["A1"] = "WorkTimeAsistem PRODE — Informe Individual"
        ws["A1"].font = titulo_font
        ws["A2"] = f"{emp_data['nombre']} — {nombre_mes} {anno}"
        ws["A2"].font = Font(name="Calibri", size=10, color="2E6DA4")
        ws["A3"] = f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws["A3"].font = Font(name="Calibri", size=8, color="888888", italic=True)

        filas = [
            ("Empleado", emp_data["nombre"]),
            ("Jornada semanal (h)", emp_data["jornada"]),
            ("Días laborables", emp_data["laborables"]),
            ("Días fichados", emp_data["fichados"]),
            ("Días con error", emp_data["errores"]),
            ("Días sin fichar", emp_data["sin_fichar"]),
            ("Horas reales", emp_data["horas_reales"]),
            ("Objetivo mensual (h)", emp_data["objetivo"]),
            ("Diferencia (h)", emp_data["diferencia"]),
            ("Horas extra (h)", emp_data["horas_extra"]),
        ]

        for idx, (campo, valor) in enumerate(filas, start=5):
            ws.cell(row=idx, column=1, value=campo).font = Font(
                name="Calibri", bold=True, size=9
            )
            val_cell = ws.cell(row=idx, column=2, value=valor)
            val_cell.font = Font(name="Calibri", size=9)
            fill_color = COLOR_FILA_IMPAR if idx % 2 == 0 else COLOR_FILA_PAR
            if campo == "Días sin fichar":
                if valor == 0:
                    fill_color = COLOR_EXITO
                elif valor < 3:
                    fill_color = COLOR_ALERTA
                else:
                    fill_color = COLOR_ERROR
            ws.cell(row=idx, column=1).fill = _fill(fill_color)
            val_cell.fill = _fill(fill_color)
            for col in [1, 2]:
                ws.cell(row=idx, column=col).border = BORDE

        buffer = BytesIO()
        wb.save(buffer)
        logger.info("Excel individual generado: %s", emp_data["nombre"])
        return buffer.getvalue()
