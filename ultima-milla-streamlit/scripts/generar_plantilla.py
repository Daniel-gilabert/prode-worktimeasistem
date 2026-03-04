import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Servicios'

def fill(hex_): return PatternFill('solid', fgColor=hex_)
def thin_border():
    s = Side(style='thin', color='FFBBBBBB')
    return Border(left=s, right=s, top=s, bottom=s)

secciones = [
    ('DATOS BÁSICOS', 'FF1F3864', [
        ('codigo',               18, 'SRV-001'),
        ('descripcion',          35, 'Reparto zona norte Valencia'),
        ('zona',                 20, 'Valencia Norte'),
        ('dimension',            18, 'Grande / Mediano'),
        ('fecha_inicio_contrato',20, '2024-01-15'),
        ('fecha_fin_contrato',   20, '(vacío = ACTIVO)'),
    ]),
    ('ASIGNACIÓN ⚠️ OBLIGATORIO', 'FFB71C1C', [
        ('empleado_base_id',     20, '5  ← ver hoja IDs'),
        ('vehiculo_base_id',     20, '16 ← ver hoja IDs'),
    ]),
    ('EMPRESA CLIENTE', 'FF2E5090', [
        ('empresa_nombre',       30, 'Logística España S.L.'),
        ('empresa_cif',          18, 'B12345678'),
        ('empresa_direccion',    35, 'Calle Mayor 45, 2ºA'),
        ('empresa_cp',           12, '46001'),
        ('empresa_ciudad',       20, 'Valencia'),
        ('empresa_provincia',    20, 'Valencia'),
        ('empresa_pais',         15, 'España'),
    ]),
    ('CONTACTO PRINCIPAL', 'FF1565C0', [
        ('contacto_nombre',      25, 'María García'),
        ('contacto_cargo',       20, 'Responsable Logística'),
        ('contacto_email',       30, 'mgarcia@empresa.com'),
        ('contacto_telefono',    18, '961234567'),
        ('contacto_movil',       18, '612345678'),
    ]),
    ('CONTACTO SECUNDARIO', 'FF6A1B9A', [
        ('contacto2_nombre',     25, 'Juan López'),
        ('contacto2_email',      30, 'jlopez@empresa.com'),
        ('contacto2_telefono',   18, '963456789'),
    ]),
    ('FACTURACIÓN', 'FF2E7D32', [
        ('facturacion_email',    30, 'facturas@empresa.com'),
        ('facturacion_forma_pago',22,'Transferencia'),
        ('numero_cuenta',        30, 'ES91 2100 0418 4502 0005 1332'),
    ]),
    ('OTROS', 'FF4E342E', [
        ('observaciones',        40, 'Horario especial en agosto'),
    ]),
]

col = 1
section_ranges = []
for sec_titulo, sec_color, campos in secciones:
    start = col
    for header, ancho, ejemplo in campos:
        c2 = ws.cell(row=2, column=col, value=header)
        c2.fill      = fill(sec_color)
        c2.font      = Font(bold=True, size=9, color='FFFFFFFF')
        c2.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c2.border    = thin_border()
        c3 = ws.cell(row=3, column=col, value=ejemplo)
        c3.fill      = fill('FFF2F2F2')
        c3.font      = Font(italic=True, size=8, color='FF777777')
        c3.alignment = Alignment(horizontal='left', vertical='center')
        c3.border    = thin_border()
        ws.column_dimensions[get_column_letter(col)].width = ancho
        col += 1
    section_ranges.append((sec_titulo, sec_color, start, col - 1))

for sec_titulo, sec_color, start, end in section_ranges:
    ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
    c1 = ws.cell(row=1, column=start, value=sec_titulo)
    c1.fill      = fill(sec_color)
    c1.font      = Font(bold=True, size=10, color='FFFFFFFF')
    c1.alignment = Alignment(horizontal='center', vertical='center')

ws.cell(row=3, column=1).value = '← EJEMPLO — borra esta fila antes de importar'
ws.row_dimensions[1].height = 20
ws.row_dimensions[2].height = 35
ws.row_dimensions[3].height = 16

total_cols = col - 1
for row in range(4, 36):
    bg = 'FFFFFFFF' if row % 2 == 0 else 'FFFFFDE7'
    for c in range(1, total_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill      = fill(bg)
        cell.border    = thin_border()
        cell.alignment = Alignment(vertical='center')

# ── Hoja IDs ─────────────────────────────────────────────────────
from supabase import create_client
sb = create_client(
    'https://drjnffoyzuploatfcltf.supabase.co',
    'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImRyam5mZm95enVwbG9hdGZjbHRmIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc3MDkzMTA5MywiZXhwIjoyMDg2NTA3MDkzfQ.dJbTqUCQh6T4H2r6ThBm-CcuzSw0diiGNKsdIDFoyiY'
)
emp = sb.table('empleados').select('id,nombre,apellidos').eq('activo', True).order('apellidos').execute().data
veh = sb.table('vehiculos').select('id,matricula,marca,modelo').neq('matricula','').order('matricula').execute().data
veh = [v for v in veh if v.get('matricula')]

wi = wb.create_sheet('IDs - Empleados y Vehículos')
wi.column_dimensions['A'].width = 10
wi.column_dimensions['B'].width = 35
wi.column_dimensions['C'].width = 10
wi.column_dimensions['D'].width = 15
wi.column_dimensions['E'].width = 20
wi.column_dimensions['F'].width = 20

# Cabeceras
for col_i, txt in enumerate(['ID_EMPLEADO','NOMBRE EMPLEADO','','ID_VEHICULO','MATRICULA','TIPO'], 1):
    c = wi.cell(row=1, column=col_i, value=txt)
    c.fill = fill('FF1F3864')
    c.font = Font(bold=True, color='FFFFFFFF', size=10)
    c.alignment = Alignment(horizontal='center')

for i, e in enumerate(emp, 2):
    wi.cell(row=i, column=1, value=e['id']).font      = Font(bold=True, color='FFCC0000')
    wi.cell(row=i, column=2, value=f"{e['apellidos']}, {e['nombre']}")

for i, v in enumerate(veh, 2):
    wi.cell(row=i, column=4, value=v['id']).font      = Font(bold=True, color='FF006600')
    wi.cell(row=i, column=5, value=v['matricula'])
    wi.cell(row=i, column=6, value=f"{v['marca']} {v['modelo']}")

wb.save('PLANTILLA_SERVICIOS.xlsx')
print('OK — PLANTILLA_SERVICIOS.xlsx regenerada con IDs y campo dimension')
